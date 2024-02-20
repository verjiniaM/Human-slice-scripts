import pandas as pd
import glob
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
import matplotlib as mpl

#%% FUNCTIONS 

def concat_dfs_in_folder(folder_path):
    '''
    reads all the excels files from the folder 
    example input(str): verji_dir + 'OP*' + '/data_tables/'  + '*RMP_high_K' + '*.xlsx'
    output: pandas datafframe with all excels concatinated

    '''
    if folder_path[-5:] == '.xlsx':
        df_paths = sorted(glob.glob(folder_path))
    else:
        df_paths = sorted(glob.glob(folder_path + '*.xlsx'))
    df_combined = pd.DataFrame()
    for path in df_paths:
        df = pd.read_excel(path)
        df_combined = pd.concat([df_combined[:], df]).reset_index(drop=True)
    return df_combined

def create_full_results_df(meta_keep_path, results_QC_path, verji_dir):

    meta_keep = pd.read_excel(meta_keep_path)

    meta_keep_IDs = []
    for i in range(len(meta_keep)):
        meta_keep_IDs.append(meta_keep['Name of recording'][i][:-4] + '_' + str(meta_keep['Channels to use'][i]))

    #results
    results_df = pd.read_excel(results_QC_path + 'complete_QC_results.xlsx')
    results_file = pd.ExcelFile(results_QC_path + 'complete_QC_results.xlsx')

    results_IDs = []
    for j in range(len(results_df)):
        results_IDs.append(results_df['Recording filename'][j][:-4] + '_' + str(results_df['Channel'][j]))

    missing_file_lines = list(set(results_file.sheet_names) - set(results_IDs))
    in_meta_not_in_results = list(set(meta_keep_IDs) - set(results_IDs))
    results_exclude = list(set(results_IDs) - set(meta_keep_IDs))

    #if all of the above 3 lists are empy continue with a piece of heart

    #take more mewtadata from the tables in the data folders of the recordings
    RMP_high_K_all = concat_dfs_in_folder(verji_dir + 'OP*' + '/data_tables/'  + '*RMP_high_K' + '*.xlsx')

    OPs, hrs_incub, cell_ID, recording_time, resting_potential, holding_minus_70_y_o_n, incubation_solution, \
        recording_in, tissue_source, patient_age, K_concentration,  temp = [],[], [], [], [], [], [], [], [], [], [], []
    #adding metadata columns 
    for i, fn in enumerate(results_df['Recording filename']):
        chan = results_df['Channel'][i]
        dat_to_add = RMP_high_K_all[(RMP_high_K_all['filename'] == fn) & 
                    (RMP_high_K_all['cell_ch'] == chan)]
        if len(dat_to_add) == 0:
            results_df =  results_df.drop(i)
            continue
        
        OPs.append(dat_to_add['OP'].item())
        hrs_incub.append(dat_to_add['hrs_incubation'].item())
        cell_ID.append(dat_to_add['cell_ID'].item())
        recording_time.append(dat_to_add['recording_time'].item())
        list_restings = dat_to_add['resting_potential'].item()[1:-1].split(',') 
        RMP = np.mean([float(i) for i in list_restings ])
        resting_potential.append(RMP)
        holding_minus_70_y_o_n.append(dat_to_add['holding_minus_70_y_o_n'].item())
        incubation_solution.append(dat_to_add['incubation_solution'].item())
        recording_in.append(dat_to_add['recording_in'].item())
        tissue_source.append(dat_to_add['tissue_source'].item())
        patient_age.append(dat_to_add['patient_age'].item())
        K_concentration.append(dat_to_add['K concentration'].item())
        temp.append(dat_to_add['temperature'].item())

    list_of_lists = [OPs, hrs_incub, cell_ID, recording_time, resting_potential, holding_minus_70_y_o_n, incubation_solution, \
        recording_in, tissue_source, patient_age, K_concentration, temp]
    list_col_names = ['OP', 'hrs_incub', 'cell_ID', 'recording_time', 'resting_potential', 'holding_minus_70_y_o_n', 'incubation_solution', \
        'recording_in', 'tissue_source', 'patient_age', 'K_concentration', 'temperature']

    #columns_to_include = list(set(list(RMP_high_K_all.columns)) - set(list(results_df.columns)))
    for i, col in enumerate(list_of_lists):
        results_df.insert(i, list_col_names[i], col)
    
    return results_df


def QC_filter_for_plotting(results_df, holding, temp, K_concentr = 8, min_age = 0):
    '''
    holding [str] - 'yes' or 'no'
    temp [str] - 'yes' or 'no'
    '''
    results_df_plot = results_df[(results_df['holding_minus_70_y_o_n'] == holding) &\
        (results_df['K_concentration'] == K_concentr) &\
            (results_df['recording_in'] != 'puff high K') &\
                #(results_df['Average amplitude (pA)'] > min_avg_amp) &\ 
                    (results_df['temperature'] == temp) &
                    (results_df['patient_age'] > min_age)]

    amp_non_negative = [-i for i in results_df_plot['Average amplitude (pA)']]
    results_df_plot.insert(20, 'Average amp (positive)', amp_non_negative)
    return results_df_plot

def QC_RMP_Ctrl(df, max_allowed_RMP_Ctrl):
    '''
    checks that the RMP in Ctrl condition is not above max_allowed_RMP_Ctrl
    exludes cells from both conditions, where it is
    '''

    #
    not_repeated_cells = []
    for cell in df['cell_ID'].unique():
        #print(len(df[df['cell_ID'] == cell]))
        if len(df[df['cell_ID'] == cell]) == 1: # or len(df[df['cell_ID'] == cell]) == 3:
            not_repeated_cells.append(cell)
    for cell in not_repeated_cells:
        df = df.drop(df.index[df['cell_ID'] == cell])     
    df.reset_index(inplace = True, drop = True)

    cell_ID_keep = df['cell_ID'][(df['recording_in'] == 'Ctrl') & (df['resting_potential'] < max_allowed_RMP_Ctrl)].to_list()
    cells_to_delete = list(set(df['cell_ID'].unique().tolist()) - set(cell_ID_keep))

    for cell in cells_to_delete:
        df = df.drop(df.index[df['cell_ID'] == cell])     
    df.reset_index(inplace = True, drop = True)
    return df

#plotting funcs
#mpl.rcParams - for all parameter settings
@mpl.rc_context({'axes.labelsize': 17, \
    'axes.spines.right': False, \
        'axes.spines.top': False, \
            'axes.titlesize': 15, \
                'xtick.labelsize': 15, \
                     'ytick.labelsize': 15, \
                        'figure.titlesize': 20})
def plot_ (df, title, params = ['Average amp (positive)', 'Average interevent interval (ms)']):
    OP_colors = ['#dede00', '#ff7f00', '#4daf4a', '#984ea3', 'violet']
    op_color_dict = {'OP230914':'#dede00', 'OP231005':'#ff7f00', 'OP231109':'#4daf4a', 'OP231123':'#984ea3', 'OP231130':'violet'}
    df = df.sort_values(by = ['recording_in', 'cell_ID']).reset_index(drop= True)
    fig, ax = plt.subplots(2,1, sharex = False, figsize=(8,10))
    for p, param in enumerate(params):
        x_vals = []
        for i, rec_solution in enumerate(sorted(df.recording_in.unique())):
            df_plot = df[df['recording_in'] == rec_solution].reset_index(drop= True) 
            x = np.linspace(2*i, 1 + 2*i, len(df_plot))
            ax[p].scatter(x, df_plot[param], alpha = 0.7, label = 'Ctrl')
            ax[p].plot([0.4 + 2*i, 0.6 + 2*i], [np.nanmean(df_plot[param]), np.nanmean(df_plot[param])], c = 'k')
            ax[p].text(0.4 + 2*i, np.nanmean(df_plot[param]) + p +0.03,  str(round(np.nanmean(df_plot[param]),2)), size = 15, c = 'k', zorder = 10)
            ax[p].set_title(param)

            x_vals.append(x)

            for j, OP in enumerate(sorted(df_plot.OP.unique())):
                indx = df_plot[df_plot['OP'] == OP].index
                x_op = x[indx]
                y_op = df_plot[param][indx]
                ax[p].scatter(x_op, y_op, c = op_color_dict[OP], s = 60, zorder = 5, label = OP)
                
        cell_IDs = df['cell_ID'][df['recording_in'] == 'Ctrl'].values
        for c, cell in enumerate(cell_IDs):
            #indx = index_s[c]
            #x_K = results_df_plot['x'].loc[results_df_plot['cell_ID'] == cell].tolist()[0]
            x = [x_vals[0][c], x_vals[1][c]]
            y = [df[param][df['recording_in'] == 'Ctrl'].tolist()[c], df[param][df['recording_in'] == 'high K'].tolist()[c]]
            #op = df_plot['OP'][df_plot['cell_ID_new'] == cell].tolist()[0]
            ax[p].plot(x, y, '-', color = 'darkgrey', alpha = 0.5, linewidth = 2, zorder = 1)
    
    ax[0].set_ylabel('Amplitude (pA)')
    ax[1].set_ylabel('IEI (ms)')

    ax[1].set_xticks(ticks = [0.5, 2.5], labels = ['Ctrl', 'high K'],)

    fig.suptitle(title)
    fig.tight_layout()
    fig.patch.set_facecolor('white')
    #fig.legend()

    plt.show()

@mpl.rc_context({'axes.labelsize': 17, \
    'axes.spines.right': False, \
        'axes.spines.top': False, \
            'axes.titlesize': 15, \
                'xtick.labelsize': 15, \
                     'ytick.labelsize': 15, \
                        'figure.titlesize': 20})
def sns_plot_MEA_data(df, title):
    colors = ['darkblue','#4daf4a', 'violet']
    customPalette = sns.set_palette(sns.color_palette(colors))

    fig, ax1 = plt.subplots(1,1, sharex = False, figsize=(8,5))
    sns.lineplot(
        data = df, x = 'Condition', y = "value",
        hue="OP", palette = customPalette , ax = ax1)
    sns.scatterplot(
        data = df, x = 'Condition', y = "value", 
        hue="OP", palette = customPalette , ax = ax1)

    ax1.set_xticks(ax1.get_xticks(), df['Condition'].unique(), rotation=30)
    
    ax1.set_xlabel('')
    ax1.set_ylabel('Network Activity \n(spikes\electrode\second')

    fig.suptitle(title)
    fig.tight_layout()
    fig.patch.set_facecolor('white')
    #fig.legend()

    plt.show()


#%%

#ANALYSIS MAIN

meta_keep_path = '/Users/verjim/laptop_D_17.01.2022/Schmitz_lab/data/human/meta_events/EPSPs/meta_dfs/meta_keep.xlsx'
results_QC_path = '/Users/verjim/laptop_D_17.01.2022/Schmitz_lab/data/human/meta_events/EPSPs/results_df/QC_ed/'
verji_dir = '/Users/verjim/laptop_D_17.01.2022/Schmitz_lab/data/human/data_verji/'

results_df = create_full_results_df(meta_keep_path, results_QC_path, verji_dir)

results_df_plot = QC_filter_for_plotting(results_df, holding = 'no', temp = 'no', min_age = 18)
results_df_plot = QC_RMP_Ctrl(results_df_plot, max_allowed_RMP_Ctrl = -50)

plot_ (results_df_plot, 'age > 18, no holding, no temp, RMP Ctrl < -50 mV')

#%%
'''
Try to see if there's a difference if Ctrl first or later on
'''
Ctrl_first = pd.read_excel('/Users/verjim/laptop_D_17.01.2022/Schmitz_lab/data/human/meta_events/EPSPs/results_df/QC_ed/Ctrl_then_high_K.xlsx')
Ctrl_plot = QC_filter_for_plotting(Ctrl_first, holding = 'no', temp = 'no')
Ctrl_plot = QC_RMP_Ctrl(Ctrl_plot, max_allowed_RMP_Ctrl = -50)

plot_ (Ctrl_plot, 'Ctrl aCSF --> High K, age > 18, no temp, RMP Ctrl < -50 mV')


high_K_first = pd.read_excel('/Users/verjim/laptop_D_17.01.2022/Schmitz_lab/data/human/meta_events/EPSPs/results_df/QC_ed/high_k_then_ctrl_aCSF.xlsx')
high_K_plot = QC_filter_for_plotting(high_K_first, holding = 'no', temp = 'no')
high_K_plot = QC_RMP_Ctrl(high_K_plot, max_allowed_RMP_Ctrl = -50)


plot_(high_K_plot, 'high K --> Ctrl aCSF, age > 18, no temp, RMP Ctrl < -50 mV')



#%%
#plottig of the MEA data

df_MEA = pd.read_excel('/Users/verjim/laptop_D_17.01.2022/Schmitz_lab/results/human/data/EPSPs_highK/MEA_Overview_MIchael_sns.xlsx')
df_MEA_Ctrl_storage = df_MEA[df_MEA['Storage'] == 'CTRL']

sns_plot_MEA_data(df_MEA_Ctrl_storage, 'Absolute change in detected spikes')


####
# OTHER FUNCTIONS THAT MIGHT BE USEFUL
# IN THE FUTURE

#%%

def plot_from_full_results_table(results_df_plot, title):
    median_amp_no_hold_no_temp_Ctrl = results_df_plot['Average amp (positive)'][results_df_plot['recording_in'] == 'Ctrl'].values
    median_amp_no_hold_no_temp_highK = results_df_plot['Average amp (positive)'][results_df_plot['recording_in'] == 'high K'].values
    freq_no_hold_no_temp_Ctrl = results_df_plot['Average interevent interval (ms)'][results_df_plot['recording_in'] == 'Ctrl'].values
    freq_no_hold_no_temp_highK = results_df_plot['Average interevent interval (ms)'][results_df_plot['recording_in'] == 'high K'].values

    fig, ax = plt.subplots(2,1, sharex = True, figsize=(12,10))
    x1 = np.linspace(0, 1,len(median_amp_no_hold_no_temp_Ctrl))
    x2 = np.linspace(2, 3,len(median_amp_no_hold_no_temp_highK))
    ax[0].scatter(x1, median_amp_no_hold_no_temp_Ctrl, alpha = 0.7, label = 'Ctrl')
    ax[0].scatter(x2, median_amp_no_hold_no_temp_highK, alpha = 0.7, label = 'high K')
    ax[0].plot([0.4, 0.6], [np.nanmean(median_amp_no_hold_no_temp_Ctrl), np.nanmean(median_amp_no_hold_no_temp_Ctrl)], c = 'k')
    ax[0].plot([2.4, 2.6], [np.nanmean(median_amp_no_hold_no_temp_highK), np.nanmean(median_amp_no_hold_no_temp_highK)], c = 'k')

    ax[0].text(0.5, np.nanmean(median_amp_no_hold_no_temp_Ctrl) + 0.07,  str(round(np.nanmean(median_amp_no_hold_no_temp_Ctrl),2)), size = 15, c = 'k')
    ax[0].text(2.5, np.nanmean(median_amp_no_hold_no_temp_highK) + 0.07,  str(round(np.nanmean(median_amp_no_hold_no_temp_highK),2)), size = 15, c = 'k')
    ax[0].set_title('Average EPSP amplitude for cell per condition, no APs')

    cell_IDs = results_df_plot['cell_ID'][results_df_plot['recording_in'] == 'Ctrl'].values
    for c, cell in enumerate(cell_IDs):
        #indx = index_s[c]
        #x_K = results_df_plot['x'].loc[results_df_plot['cell_ID'] == cell].tolist()[0]
        x = [x1[c],x2[c]]
        y = [median_amp_no_hold_no_temp_Ctrl[c], median_amp_no_hold_no_temp_highK[c]]
        #op = df_plot['OP'][df_plot['cell_ID_new'] == cell].tolist()[0]
        ax[0].plot(x, y, '-', color = 'darkgrey', alpha = 0.5, linewidth = 2, zorder = 1)

    x3 = np.linspace(0, 1,len(freq_no_hold_no_temp_Ctrl))
    x4 = np.linspace(2, 3,len(freq_no_hold_no_temp_highK))
    ax[1].scatter(x3, freq_no_hold_no_temp_Ctrl, alpha = 0.7, label = 'Ctrl')
    ax[1].scatter(x4, freq_no_hold_no_temp_highK, alpha = 0.7, label = 'high K')
    ax[1].plot([0.4, 0.6], [np.nanmean(freq_no_hold_no_temp_Ctrl), np.nanmean(freq_no_hold_no_temp_Ctrl)], c = 'k')
    ax[1].plot([2.4, 2.6], [np.nanmean(freq_no_hold_no_temp_highK), np.nanmean(freq_no_hold_no_temp_highK)], c = 'k')

    ax[1].text(0.5, np.nanmean(freq_no_hold_no_temp_Ctrl) + 0.1,  str(round(np.nanmean(freq_no_hold_no_temp_Ctrl),2)), size = 15, c = 'k')
    ax[1].text(2.5, np.nanmean(freq_no_hold_no_temp_highK) + 0.1,  str(round(np.nanmean(freq_no_hold_no_temp_highK),2)), size = 15, c = 'k')
    ax[1].set_title('Average interevent interval (ms), no APs')

    cell_IDs = results_df_plot['cell_ID'][results_df_plot['recording_in'] == 'Ctrl'].values
    for c, cell in enumerate(cell_IDs):
        #indx = index_s[c]
        #x_K = results_df_plot['x'].loc[results_df_plot['cell_ID'] == cell].tolist()[0]
        x = [x3[c],x4[c]]
        y = [freq_no_hold_no_temp_Ctrl[c], freq_no_hold_no_temp_highK[c]]
        #op = df_plot['OP'][df_plot['cell_ID_new'] == cell].tolist()[0]
        ax[1].plot(x, y, '-', color = 'darkgrey', alpha = 0.5, linewidth = 2, zorder = 1)

    ax[0].set_ylabel('Amplitude (pA)')
    ax[1].set_ylabel('IEI (ms)')

    ax[1].set_xticks(ticks = [0.5, 2.5], labels = ['Ctrl', 'high K'])

    fig.suptitle(title)
    fig.tight_layout()
    fig.patch.set_facecolor('white')
    fig.legend()

    plt.show()

#%%

# meta_dfs_dir = '/Users/verjim/laptop_D_17.01.2022/Schmitz_lab/data/human/meta_events/EPSPs/meta_dfs/original_/'
# meta_all = concat_dfs_in_folder(meta_dfs_dir) 

# #find the duplicated rows
# #duplicates_meta = meta_all.loc[meta_all.duplicated()].sort_values(['Name of recording', 'Channels to use']).reset_index(drop = True)

# meta_keep = meta_all.loc[meta_all['comment'] == 'keep'].reset_index(drop = True)

# meta_keep_IDs = []
# for i in range(len(meta_keep)):
#     meta_keep_IDs.append(meta_keep['Name of recording'][i][:-4] + '_' + str(meta_keep['Channels to use'][i]))


# results_all_path = '/Users/verjim/laptop_D_17.01.2022/Schmitz_lab/data/human/meta_events/EPSPs/results_df/complete/'
# df_name = 'results_12.01.2024_all.xlsx'

# results_file = pd.ExcelFile(results_all_path + df_name)
# results_df = pd.read_excel(results_all_path +  df_name)

# result_IDs_all = []
# for j in range(len(results_df)):
#     result_IDs_all.append(results_df['Recording filename'][j][:-4] + '_' + str(results_df['Channel'][j]))

# results_exclude = list(set(result_IDs_all) - set(meta_keep_IDs))

# if results_exclude != []:
#     for k in results_exclude:
#         fn = k[:-2] + '.abf'
#         chan = np.int64(k[-1])
#         indx = results_df.loc[(results_df['Recording filename'] == fn) & \
#             (results_df['Channel'] == chan)].index
#         results_df.drop(indx, axis=0, inplace=True)
#     results_df.reset_index(inplace = True, drop = True)

# result_IDs_keep = []
# for g in range(len(results_df)):
#     result_IDs_keep.append(results_df['Recording filename'][g][:-4] + '_' + str(results_df['Channel'][g]))

# missing_data = sorted(list(set(result_IDs_keep) - set(results_file.sheet_names)))
# print(missing_data)


# results_df.to_excel(results_all_path[:-9] + 'QC_ed/QC_' + df_name)

# #%%
# '''
# The following functions help to write or more work sheets from one excel file to another
# using pandas but not only
# '''

# results_all_path = '/Users/verjim/laptop_D_17.01.2022/Schmitz_lab/data/human/meta_events/EPSPs/results_df/complete/'
# df_name = 'results_EPSPs_21.12.2023_manualy_modified.xlsx'

# results_df_QC = pd.read_excel(results_all_path[:-9] + 'QC_ed/QC_' + df_name)

# result_IDs_QC = []
# for j in range(len(results_df_QC)):
#     result_IDs_QC.append(results_df_QC['Recording filename'][j][:-4] + '_' + str(results_df_QC['Channel'][j]))


# results_df = load_workbook(results_all_path + df_name)
# old_file = load_workbook(results_all_path[:-9] + 'QC_ed/QC_' + df_name)

# writer = pd.ExcelWriter(results_all_path[:-9] + 'QC_ed/QC_' + df_name, engine = 'openpyxl')
# writer.book = old_file

# for i in result_IDs_QC:
#     df = pd.read_excel(results_all_path +  df_name, sheet_name = i)

#     df.to_excel(writer, sheet_name = i)

# writer.close()

# #
# df_check = pd.ExcelFile(results_all_path[:-9] + 'QC_ed/QC_' + df_name)
# missing_data = sorted(list(set(result_IDs_QC) - set(df_check.sheet_names)))
# print(missing_data)
    

# #%%
# '''
# combine multyple excels wand their worksheets into one excel file
# '''
# results_QC_path = '/Users/verjim/laptop_D_17.01.2022/Schmitz_lab/data/human/meta_events/EPSPs/results_df/QC_ed/'

# df_name = 'QC_results_12.01.2024_pt3.xlsx'

# df_orig = pd.read_excel(results_QC_path + 'complete_QC_results.xlsx')
# df_to_add = pd.read_excel(results_QC_path +  df_name)
# summary_complete = pd.concat([df_orig.loc[:], df_to_add]).reset_index(drop=True)

# to_add_file = pd.ExcelFile(results_QC_path +  df_name)
# sheet_names = to_add_file.sheet_names

# backbone = load_workbook(results_QC_path + 'complete_QC_results.xlsx')

# writer = pd.ExcelWriter(results_QC_path + 'complete_QC_results.xlsx', engine = 'openpyxl')
# writer.book = backbone

# summary_complete.to_excel(writer, sheet_name = 'Sheet1')

# for i in sheet_names:
#     df = pd.read_excel(results_QC_path +  df_name, sheet_name = i)
#     df.to_excel(writer, sheet_name = i)

# writer.close()





#%%


# median_amp_no_hold_no_temp_highK = results_df['amplitude median'][(results_df['holding_minus_70_y_o_n'] == 'no') &\
#     (results_df['K_concentration'] == 8) &  (results_df['recording_in'] == 'high K')].values
# median_amp_no_hold_no_temp_Ctrl = results_df['amplitude median'][(results_df['holding_minus_70_y_o_n'] == 'no') &\
#     (results_df['K_concentration'] == 8) &  (results_df['recording_in'] == 'Ctrl')].values
# #remove APs
# median_amp_no_hold_no_temp_Ctrl = median_amp_no_hold_no_temp_Ctrl[median_amp_no_hold_no_temp_Ctrl < 10] 

# freq_no_hold_no_temp_highK = results_df['frequency'][(results_df['holding_minus_70_y_o_n'] == 'no') &\
#     (results_df['K_concentration'] == 8) &  (results_df['recording_in'] == 'high K')].values
# freq_no_hold_no_temp_Ctrl = results_df['frequency'][(results_df['holding_minus_70_y_o_n'] == 'no') &\
#     (results_df['K_concentration'] == 8) &  (results_df['recording_in'] == 'Ctrl')].values

median_amp_no_hold_no_temp_Ctrl = results_df_plot['amplitude median'][results_df['recording_in'] == 'Ctrl'].values
median_amp_no_hold_no_temp_highK = results_df_plot['amplitude median'][results_df['recording_in'] == 'high K'].values
freq_no_hold_no_temp_Ctrl = results_df_plot['frequency'][results_df['recording_in'] == 'Ctrl'].values
freq_no_hold_no_temp_highK = results_df_plot['frequency'][results_df['recording_in'] == 'high K'].values

fig, ax = plt.subplots(2,1, sharex = False, figsize=(12,10))
x1 = np.linspace(0, 1,len(median_amp_no_hold_no_temp_Ctrl))
x2 = np.linspace(2, 3,len(median_amp_no_hold_no_temp_highK))
ax[0].scatter(x1, median_amp_no_hold_no_temp_Ctrl, alpha = 0.7, label = 'Ctrl')
ax[0].scatter(x2, median_amp_no_hold_no_temp_highK, alpha = 0.7, label = 'high K')
ax[0].plot([0.4, 0.6], [np.nanmean(median_amp_no_hold_no_temp_Ctrl), np.nanmean(median_amp_no_hold_no_temp_Ctrl)], c = 'k')
ax[0].plot([2.4, 2.6], [np.nanmean(median_amp_no_hold_no_temp_highK), np.nanmean(median_amp_no_hold_no_temp_highK)], c = 'k')

ax[0].text(0.5, np.nanmean(median_amp_no_hold_no_temp_Ctrl) + 0.07,  str(round(np.nanmean(median_amp_no_hold_no_temp_Ctrl),2)), size = 15, c = 'k')
ax[0].text(2.5, np.nanmean(median_amp_no_hold_no_temp_highK) + 0.07,  str(round(np.nanmean(median_amp_no_hold_no_temp_highK),2)), size = 15, c = 'k')
ax[0].set_title('Median EPSP amplitude for cell per condition, no APs')

cell_IDs = results_df_plot['cell_ID'][results_df['recording_in'] == 'Ctrl'].values
for c, cell in enumerate(cell_IDs):
    #indx = index_s[c]
    #x_K = results_df_plot['x'].loc[results_df_plot['cell_ID'] == cell].tolist()[0]
    x = [x1[c],x2[c]]
    y = [median_amp_no_hold_no_temp_Ctrl[c], median_amp_no_hold_no_temp_highK[c]]
    #op = df_plot['OP'][df_plot['cell_ID_new'] == cell].tolist()[0]
    ax[0].plot(x, y, '-', color = 'darkgrey', alpha = 0.5, linewidth = 2, zorder = 1)

x3 = np.linspace(0, 1,len(freq_no_hold_no_temp_Ctrl))
x4 = np.linspace(2, 3,len(freq_no_hold_no_temp_highK))
ax[1].scatter(x3, freq_no_hold_no_temp_Ctrl, alpha = 0.7, label = 'Ctrl')
ax[1].scatter(x4, freq_no_hold_no_temp_highK, alpha = 0.7, label = 'high K')
ax[1].plot([0.4, 0.6], [np.nanmean(freq_no_hold_no_temp_Ctrl), np.nanmean(freq_no_hold_no_temp_Ctrl)], c = 'k')
ax[1].plot([2.4, 2.6], [np.nanmean(freq_no_hold_no_temp_highK), np.nanmean(freq_no_hold_no_temp_highK)], c = 'k')

ax[1].text(0.5, np.nanmean(freq_no_hold_no_temp_Ctrl) + 0.1,  str(round(np.nanmean(freq_no_hold_no_temp_Ctrl),2)), size = 15, c = 'k')
ax[1].text(2.5, np.nanmean(freq_no_hold_no_temp_highK) + 0.1,  str(round(np.nanmean(freq_no_hold_no_temp_highK),2)), size = 15, c = 'k')
ax[1].set_title('Median EPSP frequency for cell per condition, no APs')

cell_IDs = results_df_plot['cell_ID'][results_df['recording_in'] == 'Ctrl'].values
for c, cell in enumerate(cell_IDs):
    #indx = index_s[c]
    #x_K = results_df_plot['x'].loc[results_df_plot['cell_ID'] == cell].tolist()[0]
    x = [x3[c],x4[c]]
    y = [freq_no_hold_no_temp_Ctrl[c], freq_no_hold_no_temp_highK[c]]
    #op = df_plot['OP'][df_plot['cell_ID_new'] == cell].tolist()[0]
    ax[1].plot(x, y, '-', color = 'darkgrey', alpha = 0.5, linewidth = 2, zorder = 1)

ax[0].set_ylabel('Amplitude (mV)', fontsize = 15)
ax[1].set_ylabel('Frequency (Hz)', fontsize = 15)

ax[0].tick_params(axis='y', labelsize=15)
ax[0].tick_params(axis='x', labelsize=15)
ax[1].tick_params(axis='y', labelsize=15)
ax[1].tick_params(axis='x', labelsize=15)


fig.tight_layout()
fig.patch.set_facecolor('white')
fig.legend()

plt.show()


# #seabornb plots to see trend

# sns.lineplot(x='recording_in', y='amplitude median', data=results_df_plot, palette='viridis', size='cell_ID', sizes=list(np.ones(24)+1),legend=False, errorbar=None)
# plt.show()
# sns.lineplot(x='recording_in', y='frequency', data=results_df_plot, palette='viridis', size='cell_ID', sizes=list(np.ones(24)+1),legend=False, errorbar=None)
# plt.show()

