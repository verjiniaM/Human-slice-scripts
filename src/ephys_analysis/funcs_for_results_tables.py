import os
import pandas as pd
import ephys_analysis.funcs_sorting as sort
import numpy as np
import ephys_analysis.funcs_human_characterisation as hcf
import ephys_analysis.funcs_plotting_raw_traces as plotting_funcs
import ephys_analysis.funcs_con_screen as con_param
import datetime
import shutil
import glob
from openpyxl import load_workbook

def get_intrinsic_properties_df(human_dir, OP, tissue_source, patcher, age, inj):
    '''
    saves a pandas dataframe with intrinsic cell peoperties for OP
    '''
    work_dir, filenames, indices_dict, slice_names, pre_chans, post_chans = sort.get_OP_metadata(human_dir, OP, patcher)

    #creating a dir to save plots and data_tables (if not existing)
    dir_plots = sort.make_dir_if_not_existing (work_dir, 'plots')
    sort.make_dir_if_not_existing (work_dir, 'data_tables')

    #check if the traces dir is empty and only then plot the middle sweep for each filename
    sort.plot_trace_if_not_done(work_dir, dir_plots, filenames)

    #Correct indices if needed
    [print(key,':',value) for key, value in indices_dict.items()]
    if len(indices_dict['freq analyse']) != len(indices_dict['vc']): 
        print('Fix protocol names. Unequal number of VC and freq analyse protocols')
        input('Press enter when indices have been fixed')
        work_dir, filenames, indices_dict, slice_names, pre_chans, post_chans = sort.get_OP_metadata(human_dir, OP, patcher)
        [print(key,':',value) for key, value in indices_dict.items()]
 
    active_chans_meta = sort.get_json_meta(human_dir, OP, patcher, '_meta_active_chans.json')[0]
    cortex_out_time = sort.get_datetime_from_input(active_chans_meta['OP_time'][0])

    #creating the dataframe
    df_OP = pd.DataFrame(columns=['filename', 'slice', 'cell_ch', 'cell_ID', 'day', 'treatment', 
    'hrs_incubation', 'repatch', 'hrs_after_OP', 'Rs', 'Rin', 'resting_potential', 'max_spikes', 
    'Rheobase', 'AP_heigth', 'TH', 'max_depol', 'max_repol', 'membra_time_constant_tau', 'capacitance', 
    'Rheobse_ramp', 'AP_halfwidth'])

    for i in range(len(indices_dict['vc'])):
        vc = indices_dict['vc'][i]
        vm = indices_dict['resting'][i]
        char = indices_dict['freq analyse'][i]
        ramp = indices_dict['ramp'][i]
        slic = slice_names[vc]
        treatment = active_chans_meta['treatment'][i]
        day = 'D1'
        if slic[-2:] == 'D2':
            day = 'D2'

        filename_vc = work_dir + filenames[vc]
        filename_vm = work_dir + filenames[vm]
        filename_char = work_dir + filenames[char]
        filename_ramp = work_dir + filenames[ramp]
        #time_after_op = sort.get_time_after_OP(filename_char, cortex_out_time)
        #filename_con_screen = work_dir + filenames[indices_dict['con_screen'][i]]

        active_channels = active_chans_meta['active_chans'][i]
        
        cell_IDs = hcf.get_cell_IDs(filename_char, slic, active_channels)
        time_after_op = sort.get_time_after_OP(filename_char, cortex_out_time)
        Rs, Rin = hcf.get_access_resistance(filename_vc, active_channels)
        RMPs = hcf.get_RMP(filename_vm, active_channels)

        rheos, THs, THs_in_trace, swps = hcf.get_rheobase_from_ramp(filename_ramp, active_channels)
        params1_df = pd.DataFrame({'filename': filenames[char], 'slice' : slic, \
                                   'cell_ch': active_channels, 'hrs_after_OP' : time_after_op, \
                                    'cell_ID':cell_IDs, 'day' : day , 'treatment': treatment, \
                                        'Rs' : Rs, 'Rin': Rin, 'resting_potential': RMPs})

        charact_params  = hcf.all_chracterization_params(filename_char, active_channels, inj)
        df_char = pd.DataFrame.from_dict(charact_params)

        df_to_add = pd.concat([params1_df, df_char], axis = 1)
        df_to_add.insert(len(df_to_add.columns), 'Rheobse_ramp', rheos)
        if df_OP.empty:
            df_OP = df_to_add
        else:
            df_OP = pd.concat([df_OP.loc[:], df_to_add]).reset_index(drop=True)

        #plotting function
        plotting_funcs.plot_vc_holding (filename_vc, active_channels)
        plotting_funcs.plots_for_charact_file(filename_char, active_channels, inj)
        plotting_funcs.plot_rheobase_trace(filename_ramp, active_channels)
        #plotting_funcs.plot_connect(filename_con_screen, active_channels)
     
    tissue = pd.Series(tissue_source).repeat(len(df_OP))
    OPs = pd.Series(OP).repeat(len(df_OP))
    researcher = pd.Series(patcher).repeat(len(df_OP))
    patient_age = pd.Series(age).repeat(len(df_OP))
    series_df = pd.DataFrame({'tissue_source': tissue, 'OP': OPs, 'patcher': researcher, 'patient_age': patient_age}).reset_index(drop=True)

    df_intrinsic = pd.concat([series_df, df_OP], axis = 1)

    df_intrinsic.to_excel(work_dir + 'data_tables/' + OP + '_Intrinsic_and_synaptic_properties.xlsx', index=False) 
    #df_intrinsic.to_csv(work_dir + 'data_tables/' + OP[:-1] + '_Intrinsic_and_synaptic_properties.csv')

    print('Intrinsic properties DataFrame for  ' + OP + ' saved successfully. ' + '\n' + 'Exclude recordings if necessary.')

def get_QC_access_resistance_df (human_dir, OP, patcher):
    work_dir, filenames, indices_dict, slice_names, pre_chans, post_chans = sort.get_OP_metadata(human_dir, OP, patcher)
    
    if indices_dict['vc_end'] == []:
        return "no VC end files found; skipping"
        
    active_chans_meta = sort.get_json_meta(human_dir, OP, patcher, '_meta_active_chans.json')

    # [print(key,':',value) for key, value in indices_dict.items()]
    # if len(indices_dict['vc']) != len(indices_dict['vc_end']):
    #     print('Fix protocol names. Unequal number of VC and freq analyse protocols')
    #     active_chans_vc_end, vc_indx  = [], []
    #     for vc_end_indx in indices_dict['vc_end']:
    #         [vc_indx.append(i) for i, item in enumerate(indices_dict['vc']) if item < vc_end_indx]
    #         active_chans = active_chans_meta[0]['active_chans'][vc_indx[-1]]
    #         active_chans_vc_end.append(active_chans)
    #     active_chans_meta[0]['active_chans_vc_end'] = active_chans_vc_end
    #     #indices_dict['vc_end'] = list(input('replace missing vc_end file with "nan" '))
    # else:
    active_chans_meta[0]['active_chans_vc_end']  = active_chans_meta[0]['active_chans']

    #date_frame for quality control - change in Rin, Rs 
    df_qc = pd.DataFrame(columns=['OP','patcher', 'filename', 'slice', 'cell_ID', 'cell_ch', 'Rs_start', 'Rin_start', 'Rs_end', 'Rin_end', 
    'change_rs', 'change_rin'])

    for i in range(len(indices_dict['vc'])):
        vc = indices_dict['vc'][i]
        slic = slice_names[vc]
        filename_vc = work_dir + filenames[vc]
       
        active_channels = active_chans_meta[0]['active_chans_vc_end'][i]        
        cell_IDs = hcf.get_cell_IDs(filename_vc, slic, active_channels)

        Rs, Rin = hcf.get_access_resistance(filename_vc, active_channels)

        if indices_dict['vc_end'][i] == 'nan':
            data_to_add = pd.DataFrame({'OP':OP, 'patcher':patcher, 'filename':filenames[vc], 'slice':slic, 
            'cell_ID': cell_IDs, 'cell_ch': active_channels, 'Rs_start': Rs, 'Rin_start': Rin, 
            'Rs_end': float('nan'), 'Rin_end': float('nan'), 'chagne_rs': float('nan'), 'change_rin':float('nan')})
            df_qc = pd.concat([df_qc.loc[:], data_to_add]).reset_index(drop=True)
            continue

        vc_end = indices_dict['vc_end'][i]
        filename_vc_end = work_dir + filenames[vc_end]

        Rs, Rin = hcf.get_access_resistance(filename_vc, active_channels)
        Rs_end, Rin_end = hcf.get_access_resistance(filename_vc_end, active_channels)

        Rs_diff = [x - y for x,y in zip(Rs, Rs_end)]
        Rin_diff = [x - y for x,y in zip(Rin, Rin_end)]

        data_to_add = pd.DataFrame({'OP':OP, 'patcher':patcher, 'filename':filenames[vc], 'slice':slic, 
            'cell_ID': cell_IDs, 'cell_ch':active_channels, 'Rs_start': Rs, 'Rin_start': Rin, 
            'Rs_end': Rs_end, 'Rin_end': Rin_end, 'chagne_rs': Rs_diff, 'change_rin':Rin_diff})
        df_qc = pd.concat([df_qc.loc[:], data_to_add]).reset_index(drop=True)

    sort.to_json(work_dir, OP, '_meta_active_chans.json', active_chans_meta)
    df_qc.to_excel(work_dir + 'data_tables/' + OP + '_QC_measures_rs.xlsx', index=False) 
    remove_bad_data(OP, patcher, human_dir)
    #df_qc.to_csv(work_dir + 'data_tables/' + OP[:-1] + '_QC_measures_rs.csv')

def get_con_params_df (human_dir, OP, patcher):
    work_dir, filenames, indices_dict, slice_names, pre_chans, post_chans = sort.get_OP_metadata(human_dir, OP, patcher)
    json_meta = sort.get_json_meta(human_dir, OP, patcher, '_meta_active_chans.json')
    
    active_chans_screen, vc_indx  = [], []
    for con_indx in json_meta[1]['con_screen_file_indices']:
        [vc_indx.append(i) for i, item in enumerate(json_meta[0]['vc_files']) if item < con_indx]
        active_chans = json_meta[0]['active_chans'][vc_indx[-1]]
        active_chans_screen.append(active_chans)
    json_meta[1]['con_screen_active_chans'] = active_chans_screen

    for i, chans in enumerate(json_meta[1]['con_screen_file_indices']):
        con_screen_file = work_dir + filenames[chans]
        active_chans = json_meta[1]['con_screen_active_chans'][i]
        plotting_funcs.plot_connect(con_screen_file, active_chans)
    
    # add_pre_post_chans = input('Do you need to add pre and post channels in con_screen? (y / n)')
    # if add_pre_post_chans == 'y':
    #     pre_chans_all, post_chans_all = [], []
    #     for indx in indices_dict['con_screen']:
    #         pre_chans = [int(item) for item in input('Pre channels in ' + filenames[indx]).split()]
    #         post_chans = [int(item) for item in input('Post channels in ' + filenames[indx]).split()]
    #         pre_chans_all.append(pre_chans)
    #         post_chans_all.append(post_chans)
        
        # json_meta[1]['pre_chans'] = pre_chans_all
        # json_meta[1]['post_chans'] = post_chans_all
    sort.to_json(work_dir, OP, '_meta_active_chans.json', json_meta)

    con_sccreen_connected_chans = sort.get_json_meta(human_dir, OP, patcher, '_meta_active_chans.json')[1]
    cortex_out_time = sort.get_datetime_from_input(con_sccreen_connected_chans['OP_time'][0])

    con_data = pd.DataFrame(columns = ['OP', 'fn', 'slice', 'day', 'connection_ID', 'repatch', 'treatment', 
    'hrs_after_OP', 'hrs_incubation','chan_pre', 'chan_post', 'Vm pre', 'Vm post', 
    'Amp 1','Amp 2','Amp 3', 'Amp 4',	'Lat1',	'Lat2',	'Lat3',	'Lat4', 'num excluded swps', 'comments'])

    for i, indx in enumerate(con_sccreen_connected_chans['con_screen_file_indices']):
        con_screen_file = work_dir + filenames[indx]
        
        pre_cells = con_sccreen_connected_chans['pre_chans'][i]
        post_cells = con_sccreen_connected_chans['post_chans'][i]
        treatment = con_sccreen_connected_chans['treatment'][i]
        time_after_op = sort.get_time_after_OP(con_screen_file, cortex_out_time)
        
        exclude = ''
        for j, pre_cell in enumerate(pre_cells):
            post_cell = post_cells[j]
            slic = slice_names[indx]
            con_ID = hcf.get_connection_ID(con_screen_file, slic, pre_cell, post_cell)
            day = 'D1'
            if slic[-2:] == 'D2':
                day = 'D2'

            pre_signal, es, vm_pre = con_param.presynaptic_screen(con_screen_file, pre_cell)
            post_signal, vm_post = con_param.postsynaptic_screen(con_screen_file, post_cell, es)
            if (np.array(vm_post)).size == 0:
                print('QC not passed!!')
                exclude = 'all'
                es2 = []
                post_signal, vm_post = con_param.postsynaptic_screen(con_screen_file, post_cell, es2)
                continue
            mean_pre, mean_post, pre_window, post_window, preAPs_shifted, preAPs = \
                con_param.get_analysis_window(pre_signal, post_signal)
            pre_signal, post_signal = con_param.remove_sweeps_with_POSTsynAPs(pre_signal, post_signal, preAPs)
            post_peaks = con_param.find_postsynaptic_peaks(post_window, preAPs_shifted)
            post_local_baseline = con_param.get_psp_baselines(post_window,preAPs_shifted)
            onsets = con_param.get_onsets(preAPs_shifted, post_window, post_peaks, post_local_baseline)
            latency = con_param.latencies(onsets, preAPs_shifted)
            amps = con_param.get_amps(post_peaks, post_local_baseline)

            df_add = pd.DataFrame({'OP': OP, 'fn': filenames[indx], 'slice': slic, 'day':day, 'treatment': treatment, 
            'connection_ID' : con_ID, 'hrs_after_OP' : time_after_op,
            'chan_pre': pre_cell, 'chan_post': post_cell, 'Vm pre' :vm_pre, 'Vm post': vm_post,
            'Amp 1': amps[0], 'Amp 2': amps[1],	'Amp 3': amps[2],	'Amp 4': amps[3],	
             'Lat1': latency[0][0],	'Lat2' : latency[1][0],	'Lat3': latency[2][0],	'Lat4': latency[3][0], 
            'num excluded swps': len(es), 'comments': exclude}, index=[0])
            con_data = pd.concat([con_data.loc[:], df_add]).reset_index(drop=True)

            #plotting
            plotting_funcs.plot_connection_window(con_screen_file, pre_cell, post_cell, pre_window, \
                post_window, preAPs_shifted, post_signal, onsets, preAPs, post_peaks, post_local_baseline)
            plotting_funcs.plot_post_cell(con_screen_file, pre_cell, post_cell)
            #plotting_funcs.plot_post_cell_old_win(con_screen_file, pre_cell, post_cell)

    con_data.to_excel(work_dir + '/data_tables/' + OP + '_connected_cell_properties.xlsx', index=False) 

def get_con_screen_VC (human_dir, OP, patcher):
    work_dir, filenames, indices_dict, slice_names, pre_chans, post_chans = sort.get_OP_metadata(human_dir, OP, patcher)
    json_meta = sort.get_json_meta(human_dir, OP, patcher, '_meta_active_chans.json')
    
    add_pre_post_chans = input('Do you need to add pre and post channels in con_screen_VC? ( y / n)')
    if add_pre_post_chans == 'y':
        pre_chans_IC, post_chans_VC = [], []
        for con_indx in json_meta[4]['con_screen_IC_file_indices']:
            pre_chans_IC = [int(item) for item in input('Pre channels in IC in ' + filenames[indx]).split()]
            post_chans_VC = [int(item) for item in input('Post channels in  VC in ' + filenames[indx]).split()]
            pre_chans_IC.append(pre_chans_IC)
            post_chans_VC.append(post_chans_VC)
        json_meta[4]['pre_chans_IC'] = pre_chans_IC
        json_meta[4]['post_chans_VC'] = post_chans_VC
        sort.to_json(work_dir, OP, '_meta_active_chans.json', json_meta)

    con_sccreen_connected_chans = sort.get_json_meta(human_dir, OP, patcher, '_meta_active_chans.json')[4]
    cortex_out_time = sort.get_datetime_from_input(con_sccreen_connected_chans['OP_time'][0])

    con_data = pd.DataFrame(columns = ['OP', 'fn', 'slice', 'day', 'connection_ID', 'repatch', 'treatment', 
    'hrs_after_OP', 'hrs_incubation','chan_pre', 'chan_post', 'Vm pre', 'Holding post', 
    'Amp 1',	'Amp 2',	'Amp 3',	'Amp 4',	'Lat1',	'Lat2',	'Lat3',	'Lat4', 'num excluded swps', 'comments'])

    for i, indx in enumerate(con_sccreen_connected_chans['con_screen_IC_file_indices']):
        con_screen_file_IC = work_dir + filenames[indx]
        
        pre_cells = con_sccreen_connected_chans['pre_chans_IC'][i]
        post_cells = con_sccreen_connected_chans['post_chans_VC'][i]
        if pre_cells == [] or post_cells == []:
            continue
        treatment = con_sccreen_connected_chans['treatment'][i]
        time_after_op = sort.get_time_after_OP(con_screen_file_IC, cortex_out_time)
        
        exclude = ''
        for j, pre_cell in enumerate(pre_cells):
            post_cell = post_cells[j]
            slic = slice_names[indx]
            con_ID = hcf.get_connection_ID (con_screen_file_IC, slic, pre_cell, post_cell)

            day = 'D1'
            if slic[-2:] == 'D2':
                day = 'D2'

            pre_sig, es, vm_pre = con_param.presynaptic_screen_IC(con_screen_file_IC, pre_cell)
            post_sig, holding_post = con_param.postsynaptic_screen_VC (con_screen_file_IC, post_cell, es)
            if (np.array(holding_post)).size == 0:
                print('QC not passed!!')
                exclude = 'all'
                es2 = []
                post_sig, holding_post = con_param.postsynaptic_screen_VC (con_screen_file_IC, post_cell, es)
                continue
            mean_pre, mean_post, pre_window, post_window, preAPs_shifted, preAPs = con_param.get_analysis_window_VC(pre_sig, post_sig)
            post_peaks = con_param.find_postsynaptic_peaks_VC(post_window, preAPs_shifted)
            bl = con_param.get_psp_baselines(post_window,preAPs_shifted)
            onsets = con_param.get_onsets_VC(preAPs_shifted, post_window, post_peaks, bl)
            latency = con_param.latencies(onsets, preAPs_shifted)
            amps = con_param.get_amps_VC(post_peaks, bl)

            df_add = pd.DataFrame({'OP': OP, 'fn': filenames[indx], 'slice': slic, 'day':day, 'treatment': treatment, 
            'connection_ID' : con_ID, 'hrs_after_OP' : time_after_op,
            'chan_pre': pre_cell, 'chan_post': post_cell, 'Vm pre' :vm_pre, 'Holding post': holding_post,
            'Amp 1': amps[0], 'Amp 2': amps[1],	'Amp 3': amps[2],	'Amp 4': amps[3],	
             'Lat1': latency[0][0],	'Lat2' : latency[1][0],	'Lat3': latency[2][0],	'Lat4': latency[3][0], 
            'num excluded swps': len(es), 'comments': exclude}, index=[0])
            con_data = pd.concat([con_data.loc[:], df_add]).reset_index(drop=True)

            #plotting
            plotting_funcs.plot_connection_window_VC(con_screen_file_IC, pre_cell, post_cell, pre_window, \
                    post_window, preAPs_shifted, post_sig, onsets, preAPs, post_peaks, bl)
            plotting_funcs.plot_post_cell_VC(con_screen_file_IC, pre_cell, post_cell)

    con_data.to_excel(work_dir + '/data_tables/' + OP + '_connected_cell_properties_post_in_VC.xlsx', index=False) 

def get_spontan_QC(human_dir, OP, patcher):
    work_dir, filenames, indices_dict, slice_names, pre_chans, post_chans = sort.get_OP_metadata(human_dir, OP, patcher)
    
    active_chans_meta = sort.get_json_meta(human_dir, OP, patcher, '_meta_active_chans.json')
    if len(indices_dict['spontan']) == len(indices_dict['vc']):
        active_chans_meta[0]['active_chans_spontan'] = active_chans_meta[0]['active_chans']
    else:
        print('Adding active chans spontan')
        active_chans_spontan, vc_indx  = [], []
        for spontan_indx in indices_dict['spontan']:
            [vc_indx.append(i) for i, item in enumerate(active_chans_meta[0]['vc_files']) if item < spontan_indx]
            active_chans = active_chans_meta[0]['active_chans'][vc_indx[-1]]
            active_chans_spontan.append(active_chans)
        active_chans_meta[0]['active_chans_spontan'] = active_chans_spontan
        sort.to_json(work_dir, OP, '_meta_active_chans.json', active_chans_meta)

    input('Are the active_chans_spontan correct?')

    #record_sorting = pd.DataFrame(columns = ['OP', 'patcher', 'filename', 'cell_ch','swps_to_analyse', 'swps_to_discard', 
    #'min_vals_each_swp', 'max_vals_each_swp', 'drift'])

    record_sorting = pd.DataFrame()
    for i, u in enumerate(indices_dict['spontan']):
        filename_spontan = work_dir + filenames[u]
        slic = slice_names[u]   

        #index_chans = active_chans_meta[0]['slices'].index(slic)
        active_channels = active_chans_meta[0]['active_chans_spontan'][i]
        cell_IDs = hcf.get_cell_IDs(filename_spontan, slic, active_channels)
        treatment = active_chans_meta[3][slic[:2]]
        #treatment = active_chans_meta[0]['treatment'][i]

        spontan_QC = hcf.rec_stability(filename_spontan, active_channels , 60)
        df_QC = pd.DataFrame(spontan_QC).T
        df_QC['cell_ch'] = df_QC.index 
        df_QC.insert(0, 'cell_ch', df_QC.pop('cell_ch'))
        df_QC.insert(0, 'slice', slic)
        df_QC.insert(0, 'cell_ID', cell_IDs)
        df_QC.insert(0, 'filename', filenames[u])
        df_QC.insert(0, 'patcher', patcher)
        df_QC.insert(0, 'OP', OP)
        df_QC.insert(len(df_QC),'treatment', treatment)

        record_sorting = pd.concat([record_sorting.loc[:], df_QC]).reset_index(drop=True)
    #remove cells where no swps to analyse
    if len(record_sorting) > 0:
        mask = []
        for i in record_sorting.swps_to_analyse:
            if len(i) <= 0:
                mask.append(False)
                continue
            mask.append(True)
        record_sorting = record_sorting.loc[mask,:]
    sort.to_json(work_dir, OP, '_meta_active_chans.json', active_chans_meta)
    record_sorting.to_excel(work_dir + 'data_tables/' + OP + '_QC_measures_spontan.xlsx', index=False) 
    #record_sorting.to_csv(work_dir + 'data_tables/' + OP + '_QC_measures_spontan.csv')

def get_minis_QC(human_dir, OP, patcher):
    work_dir, filenames, indices_dict, slice_names, pre_chans, post_chans = sort.get_OP_metadata(human_dir, OP, patcher)

    #indices_dict = sort.from_json(work_dir, OP, '_indices_dict.json')[0]
    active_chans_meta = sort.get_json_meta(human_dir, OP, patcher, '_meta_active_chans.json')
    cortex_out_time = sort.get_datetime_from_input(active_chans_meta[2]['OP_time'][0])

    record_sorting = pd.DataFrame(columns = ['OP','patcher', 'filename', 'cell_ch','hrs_incubation','swps_to_analyse', 'swps_to_discard', 
    'min_vals_each_swp', 'max_vals_each_swp', 'drift', 'Rs_change', 'Rin_change', 'Rs_start'])

    for i, indx in enumerate(indices_dict['minis']):
        mini_file = work_dir + filenames[indx]
        slic = slice_names[indx]

        #index_chans = active_chans_meta[2]['mini_slices'].index(slic)
        active_channels = active_chans_meta[2]['mini_chans'][i]
        cell_IDs = hcf.get_cell_IDs(mini_file, slic, active_channels)
        time_after_op = sort.get_time_after_OP(mini_file, cortex_out_time)
        mini_QC = hcf.rec_stability(mini_file, active_channels , 60)
        treatment = active_chans_meta[2]['treatment'][i]

        #if len(indices_dict['vc_mini']) == len(indices_dict['minis']):
        vc_indx = indices_dict['vc_mini'][i]
        Rs, Rin = hcf.get_access_resistance(work_dir + filenames[vc_indx], active_channels)
        
        #if len(indices_dict['vc_mini_end']) == len(indices_dict['minis']):
        vc_end_indx = indices_dict['vc_mini_end'][i]
        Rs_end, Rin_end = hcf.get_access_resistance(work_dir + filenames[vc_end_indx], active_channels)
        
        Rs_diff = [x - y for x,y in zip(Rs, Rs_end)]
        Rin_diff = [x - y for x,y in zip(Rin, Rin_end)]

        df_QC = pd.DataFrame(mini_QC).T
        df_QC['cell_ch'] = df_QC.index 
        df_QC.insert(0, 'cell_ch', df_QC.pop('cell_ch'))
        df_QC.insert(0, 'slice', slic)
        df_QC.insert(0, 'cell_ID', cell_IDs)
        df_QC.insert(0, 'filename', filenames[indx])
        df_QC.insert(0, 'patcher', patcher)
        df_QC.insert(0, 'OP', OP)
        df_QC.insert(len(df_QC),'treatment', treatment)
        df_QC['Rs_change'], df_QC['Rin_change'], df_QC['Rs_start'] = Rs_diff, Rin_diff, Rs

        record_sorting = pd.concat([record_sorting.loc[:], df_QC]).reset_index(drop=True)
    record_sorting.to_excel(work_dir + 'data_tables/' + OP + '_QC_measures_minis.xlsx', index=False) 

def get_intrinsic_properties_df_no_VM_file (human_dir, OP, tissue_source, patcher, age, inj):
    '''
    saves a pandas dataframe with intrinsic cell peoperties for OP
    '''
    work_dir, filenames, indices_dict, slice_names, pre_chans, post_chans = sort.get_OP_metadata(human_dir, OP, patcher)
   
    #creating a dir to save plots and data_tables (if not existing)
    dir_plots = sort.make_dir_if_not_existing (work_dir, 'plots')
    sort.make_dir_if_not_existing (work_dir, 'data_tables')

    #check if the traces dir is empty and only then plot the mimiddle sweep for each filename
    sort.plot_trace_if_not_done(work_dir, dir_plots, filenames)

    #Correct indices if needed
    [print(key,':',value) for key, value in indices_dict.items()]
    if len(indices_dict['freq analyse']) != len(indices_dict['vc']): 
        print('Fix protocol names. Unequal number of VC and freq analyse protocols')
        input('Press enter when indices have been fixed')
        work_dir, filenames, indices_dict, slice_names, pre_chans, post_chans = sort.get_OP_metadata(human_dir, OP, patcher)
        [print(key,':',value) for key, value in indices_dict.items()]
        # index_vc_in = [int(item) for item in input('Vc files corresponding to characterization files for ' + OP +' (input with spaces in between)').split()]
        # #saved the original indices
        # indices_dict['vc_orig'] = indices_dict['vc']
        # indices_dict['vc'] = index_vc_in
        # # index_char = [int(item) for item in input('corresponding characterization files for ' + OP +' (input with spaces in between)').split()]
        # # indices_dict['freq analyse'] = index_char
 
    active_chans_meta = sort.get_json_meta(human_dir, OP, patcher, '_meta_active_chans.json')[0]
    cortex_out_time = sort.get_datetime_from_input(active_chans_meta['OP_time'][0])

    #creating the dataframe
    df_OP = pd.DataFrame(columns=['filename', 'slice', 'cell_ch', 'cell_ID', 'day', 'treatment', 
    'hrs_incubation', 'repatch', 'hrs_after_OP', 'Rs', 'Rin', 'resting_potential', 'max_spikes', 'Rheobase', 
    'AP_heigth', 'TH', 'max_depol', 'max_repol', 'membra_time_constant_tau', 'capacitance', 'AP_halfwidth'])

    for i in range(len(indices_dict['vc'])):
        vc = indices_dict['vc'][i]
        char = indices_dict['freq analyse'][i]
        slic = slice_names[vc]
        day = 'D1'
        if slic[-2:] == 'D2': 
            day = 'D2'
        treatment = active_chans_meta['treatment'][i]

        filename_vc = work_dir + filenames[vc]
        filename_char = work_dir + filenames[char]
        #filename_con_screen = work_dir + filenames[indices_dict['con_screen'][i]]
        time_after_op = sort.get_time_after_OP(filename_char, cortex_out_time)

        active_channels = active_chans_meta['active_chans'][i]
        
        cell_IDs = hcf.get_cell_IDs(filename_char, slic, active_channels)
        Rs, Rin = hcf.get_access_resistance(filename_vc, active_channels) 
    
        charact_dict = hcf.load_traces(filename_char)
        inj = hcf.read_inj('full')
        tau_all, capacitance_all, mc_all, V65_all, RMPs = hcf.get_hyperpolar_param(charact_dict, active_channels, inj)
        
        params1_df = pd.DataFrame({'filename': filenames[char], 'slice' : slic, 'cell_ch': active_channels,
        'hrs_after_OP' : time_after_op,
        'cell_ID': cell_IDs, 'day' : day , 'treatment': treatment, 'Rs' : Rs, 'Rin': Rin, 'resting_potential': RMPs })        

        charact_params  = hcf.all_chracterization_params(filename_char, active_channels, inj)
        df_char = pd.DataFrame.from_dict(charact_params)

        df_to_add = pd.concat([params1_df, df_char], axis = 1)
        df_OP = pd.concat([df_OP.loc[:], df_to_add]).reset_index(drop=True)

        #plotting function
        plotting_funcs.plot_vc_holding (filename_vc, active_channels)
        plotting_funcs.plots_for_charact_file(filename_char, active_channels, inj)
        #plotting_funcs.plot_connect(filename_con_screen, active_channels)
        #plotting_funcs.plot_connect_old_win(filename_con_screen, active_channels)
     
    tissue = pd.Series(tissue_source).repeat(len(df_OP))
    OPs = pd.Series(OP).repeat(len(df_OP))
    researcher = pd.Series(patcher).repeat(len(df_OP))
    patient_age = pd.Series(age).repeat(len(df_OP))
    series_df = pd.DataFrame({'tissue_source': tissue, 'OP': OPs, 'patcher': researcher, 'patient_age': patient_age}).reset_index(drop=True)

    df_intrinsic = pd.concat([series_df, df_OP], axis = 1)

    df_intrinsic.to_excel(work_dir + 'data_tables/' + OP + '_Intrinsic_and_synaptic_properties.xlsx', index=False) 
    #df_intrinsic.to_csv(work_dir + 'data_tables/' + OP + '_Intrinsic_and_synaptic_properties.csv')

    print('Intrinsic properties DataFrame for  ' + OP + ' saved successfully. ' + '\n' + 'Exclude recordings if necessary.')


def remove_bad_data (OP, patcher, human_dir = '/Users/verjim/laptop_D_17.01.2022/Schmitz_lab/data/human/'):
    work_dir, filenames, indices_dict, slice_names, pre_chans, post_chans = sort.get_OP_metadata(human_dir, OP, patcher)

    intr_df = pd.read_excel(work_dir + 'data_tables/' + OP + '_Intrinsic_and_synaptic_properties.xlsx')
    df_vc_qc =  pd.read_excel(work_dir + 'data_tables/' + OP + '_QC_measures_rs.xlsx')

    comments = list(range(0, len(intr_df)))
    big_change_in_Rs = df_vc_qc['cell_ID'][abs(df_vc_qc['change_rs']) > 15].tolist()
    for i in range(len(big_change_in_Rs)):
        for cell in range(len(intr_df.index[intr_df['cell_ID'] == big_change_in_Rs[i]])):
            cell_indx = intr_df.index[intr_df['cell_ID'] == big_change_in_Rs[i]][cell]
            comments[cell_indx] = 'exclude; big change in Rs from start to end'

    #remove cells where Rs > 30
    Rs_QC_fail = df_vc_qc['cell_ID'][df_vc_qc['Rs_end'] > 30].tolist()
    for j in range(len(Rs_QC_fail)):
        for cell in range(len(intr_df.index[intr_df['cell_ID'] == Rs_QC_fail[j]])):
            cell_indx = intr_df.index[intr_df['cell_ID'] == Rs_QC_fail[j]][cell]
            comments[cell_indx] = str(comments[cell_indx]) + '; exclude; Rs_end > 30'

    intr_df.insert(len(intr_df.columns), 'comments', comments)

    intr_df.to_excel(work_dir + 'data_tables/' + 'QC_passed_' + OP + '_Intrinsic_and_synaptic_properties.xlsx', index=False) 

def check_cell_IDs (human_dir, OP, patcher):
    work_dir = sort.get_work_dir(human_dir, OP, patcher)
    df_intrinsic_final = pd.read_excel(work_dir + 'data_tables/QC_passed_' + OP + '_Intrinsic_and_synaptic_properties.xlsx')
    df_spontan = pd.read_excel(work_dir + 'data_tables/' + OP + '_QC_measures_spontan.xlsx')

    missing_cell_ids = []
    [missing_cell_ids.append(cell_id) for cell_id in df_intrinsic_final['cell_ID'].tolist() if cell_id not in df_spontan['cell_ID'].tolist()]

    print('Missing cell ID(s): ' + str(missing_cell_ids) + '. Found in Intrinsic props but not in spontan')

    missing_cell_ids2 = []
    [missing_cell_ids2.append(cell_id) for cell_id in df_spontan['cell_ID'].tolist() if cell_id not in df_intrinsic_final['cell_ID'].tolist()]

    print('Missing cell ID(s): ' + str(missing_cell_ids2) + '. Found in spontan but not in intrinsic props')



# following functions for data organization adn collection preceeding automatic event analysis
def get_metadata_for_event_analysis(human_dir, OP, patcher:str, event_type): # add event_type 
    '''
    event_type = 'minis', 'spontan', 'EPSPs'
    '''
    work_dir = sort.get_work_dir(human_dir, OP, patcher)

    if event_type == 'minis':
        df_events = pd.read_excel(work_dir + 'data_tables/' + OP + '_QC_measures_minis.xlsx')
        df_meta = pd.DataFrame({
            'Name of recording': df_events['filename'],  
            'Channels to use': df_events['cell_ch'], 
            'Sampling rate (Hz)': 20_000,
            'Analysis start at sweep (number)' : 1,
            'Cut sweeps first part (ms)' : 5_500,
            'Cut sweeps last part (ms)' : 0,
            'Analysis length (min)' : 2,
            'swps_to_analyse' : df_events['swps_to_analyse'],
            'OP': df_events['OP'],
            'event_type' : event_type, 
            'treatment' : df_events['treatment'], 
            'hrs_incubation' : df_events['hrs_incubation'],
            'slice' : df_events['slice'],
            'patcher' : patcher
            })

    elif event_type == 'spontan':
        df_events = pd.read_excel(work_dir + 'data_tables/' + OP + '_QC_measures_spontan.xlsx')
        df_intrinsic = pd.read_excel(glob.glob(work_dir + 'data_tables/' + 'QC_passed_' + '*.xlsx')[0])

        df_meta = pd.DataFrame({
            'Name of recording': df_events['filename'],  
            'Channels to use': df_events['cell_ch'], 
            'Sampling rate (Hz)': 20_000,
            'Analysis start at sweep (number)' : 1,
            'Cut sweeps first part (ms)' : 5_500,
            'Cut sweeps last part (ms)' : 0,
            'Analysis length (min)' : 2,
            'swps_to_analyse' : df_events['swps_to_analyse'],
            'OP': df_events['OP'],
            'event_type' : event_type,
            'cell_ID' : df_events['cell_ID'],
            'hrs_incubation': df_intrinsic['hrs_incubation'],
            'treatment' : df_events['treatment'], 
            'slice' : df_events['slice'],
            'patcher' : patcher
            })

    elif event_type == 'EPSPs':
        df_events = pd.read_excel(work_dir + 'data_tables/RMP_high_K_' + OP + '.xlsx')
        #df_intrinsic = pd.read_excel(glob.glob(work_dir + 'data_tables/' + 'QC_passed_' + '*.xlsx')[0])
        df_meta = pd.DataFrame({
            'Name of recording': df_events['filename'],
            'Channels to use': df_events['cell_ch'], 
            'Sampling rate (Hz)': 20_000,
            'Analysis start at sweep (number)' : 1,
            'Cut from datapoint' : 0,
            'Cut to datapoint' : 0,
            'swps_to_delete' : None,
            'OP': df_events['OP'],
            'event_type' : event_type, 
            'recording_in' : df_events['recording_in'], 
            'slice' : df_events['slice'],
            'patcher' : patcher
            })
    else: 
        print('Please enter a valid event type')

    df_meta.to_excel(work_dir + 'data_tables/' + event_type + '_meta_' + OP + '.xlsx', index=False) 


# Functions for full analysis (not OP-based)

def prapare_for_event_analysis(human_dir = '/Users/verjim/laptop_D_17.01.2022/Schmitz_lab/data/human/'):
    '''
    checks mini or spontan experiments surgery in the '*experiments_overview.xlsx'
    puts all files to be analyzed in human_dir + '/meta_events/[spontan/mini]_files/'
    saves the meta_tables to be analyzed in human_dir + '/meta_events/meta_files_to_analyse/'
    '''
    #exp_view = pd.read_excel('/Users/verjim/laptop_D_17.01.2022/Schmitz_lab/data/human/2024-05-11_experiments_overview.xlsx')
    exp_view = pd.read_excel(glob.glob(human_dir + '*experiments_overview.xlsx')[0]) 
    date = str(datetime.date.today())
    
    #op_to_analyse = ['OP230914', 'OP231005', 'OP231109', 'OP231123', 'OP231130']
    meta_df_mini, meta_df_spontan, meta_df_EPSPs = pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    for i in range(len(exp_view)): #ran to range 22
        if exp_view['minis'][i] == 'yes':    
            patcher =  exp_view['patcher'][i]
            OP = exp_view['OP'][i]
            get_metadata_for_event_analysis(human_dir, OP, patcher, 'minis') 
            work_dir_mini = sort.get_work_dir(human_dir, OP, patcher)
            df_mini = pd.read_excel(work_dir_mini + 'data_tables/minis_meta_' + OP + '.xlsx') 
            df_mini = df_mini.loc[~df_mini['Name of recording'].isna()]
            #moving files directly to the recordings folder in event analysis
            for f in df_mini['Name of recording']:
                shutil.copy(os.path.join(work_dir_mini, f), '/Users/verjim/spontaneous-postsynaptic-currents-detection/recordings/')
            #     shutil.copy(os.path.join(work_dir_mini, f), human_dir + '/meta_events/mini_files/')
            meta_df_mini = pd.concat([meta_df_mini.loc[:], df_mini]).reset_index(drop=True)

        if exp_view['spontaneous'][i] == 'yes':    
            patcher =  exp_view['patcher'][i]
            OP = exp_view['OP'][i]
            get_metadata_for_event_analysis(human_dir, OP, patcher, 'spontan') 
            work_dir_spontan = sort.get_work_dir(human_dir, OP, patcher)
            df_spontan = pd.read_excel(work_dir_spontan + 'data_tables/spontan_meta_' + OP + '.xlsx') 
            df_spontan = df_spontan.loc[~df_spontan['Name of recording'].isna()]
            #moving files directly to the recordings folder in event analysis
            for f in df_spontan['Name of recording']:
                shutil.copy(os.path.join(work_dir_spontan, f), '/Users/verjim/spontaneous-postsynaptic-currents-detection/recordings/')
            meta_df_spontan = pd.concat([meta_df_spontan.loc[:], df_spontan]).reset_index(drop=True)

        # if exp_view['EPSPs_highK'][i] == 'yes':
        #     patcher = exp_view['patcher'][i]
        #     OP = exp_view['OP'][i]
        #     get_metadata_for_event_analysis(human_dir, OP, patcher, 'EPSPs') 
        #     work_dir_EPSPs = sort.get_work_dir(human_dir, OP, patcher)
        #     df_EPSPs = pd.read_excel(work_dir_EPSPs + 'data_tables/EPSPs_meta_' + OP + '.xlsx') 
        #     meta_df_EPSPs = pd.concat([meta_df_EPSPs.loc[:], df_EPSPs]).reset_index(drop=True)

    #remove any files where there's nothing to analyse
    meta_df_mini.reset_index(inplace = True, drop = True)
    for i in reversed(range(len(meta_df_mini))):
        if meta_df_mini['swps_to_analyse'][i] == '[]':
            meta_df_mini = meta_df_mini.drop([meta_df_mini.index[i]])

    meta_df_spontan.reset_index(inplace = True, drop = True)
    for i in reversed(range(len(meta_df_spontan))):
        if meta_df_spontan['swps_to_analyse'][i] == '[]':
            meta_df_spontan = meta_df_spontan.drop([meta_df_spontan.index[i]])
    
    # meta_df_EPSPs.reset_index(inplace = True, drop = True)
    # for i in reversed(range(len(meta_df_EPSPs))):
    #     if meta_df_EPSPs['swps_to_analyse'][i] == '[]':
    #         meta_df_EPSPs = meta_df_EPSPs.drop([meta_df_EPSPs.index[i]])

    #save data 
    meta_df_mini.to_excel(human_dir + '/meta_events/meta_files_to_analyse/' + date + 'minis_meta.xlsx',index=False)
    meta_df_spontan.to_excel(human_dir + '/meta_events/meta_files_to_analyse/' + date + 'spontan_meta.xlsx',index=False)
    #meta_df_EPSPs.to_excel(human_dir + '/meta_events/EPSPs/meta_dfs/' + date + 'EPSPs_meta.xlsx',index=False)
    print('Meta data for events is saved. Files are copied to desired folders. Ready to proceed with event analysis. ')
    return meta_df_mini, meta_df_spontan

#analysis of intrinsic properties
def collect_intrinsic_df(human_dir = '/Users/verjim/laptop_D_17.01.2022/Schmitz_lab/data/human/'):
    '''
    collects *QC_passed data tables from all OP folders
    adds area and high K concentratioon columns from exp_view
    saves a datafram with today's data in /results/human/data/summary_data_tables/intrinsic_properties/
    '''

    exp_view = pd.read_excel(glob.glob(human_dir + '*experiments_overview.xlsx')[0]) 

    area_dict, high_k_dict = {}, {}
    for OP in exp_view['OP'].unique():
        area_dict[OP] = exp_view['region'][exp_view['OP'] == OP].tolist()[0]
        high_k_dict[OP] = exp_view['K concentration'][exp_view['OP'] == OP].tolist()[0]
    
    intr_props_dirs = glob.glob(human_dir + '/data_*/' + 'OP*' + '/data_tables/' + 'QC_passed' + '*.xlsx')
    for i, intr_path in enumerate(intr_props_dirs):
        df = pd.read_excel(intr_path)
        if df.empty or df.isna().all().all():
            continue
        if i == 0:
            all_intr_props = df
            continue
        all_intr_props = pd.concat([all_intr_props.loc[:], df]).reset_index(drop=True)
     
    all_intr_props.insert(0, 'area', '')
    for OP in all_intr_props['OP'].unique():
        mask = all_intr_props['OP'] == OP
        all_intr_props.loc[mask, 'area'] = area_dict[OP]
        all_intr_props.loc[mask, 'high K concentration'] = high_k_dict[OP]

    all_intr_props.loc[all_intr_props['treatment'] == 'high k'] = 'high K'

    # date = str(datetime.date.today())
    # all_intr_props.to_excel('/Users/verjim/laptop_D_17.01.2022/Schmitz_lab/results/human/data/summary_data_tables/intrinsic_properties/' 
    # + date + '_collected.xlsx', index=False)
    return all_intr_props

def save_intr_data(df_intr_props,  QC_data, repatch_data, juv_repatch, adult_repatch):
    ''' 
    saves the collected intrinsic data properties in the general data folder
    '''
    file_name = '/Users/verjim/laptop_D_17.01.2022/Schmitz_lab/data/all_data_human.xlsx'
    with pd.ExcelWriter(file_name) as writer:  
        df_intr_props.to_excel(writer, sheet_name='all_data')
        QC_data.to_excel(writer, sheet_name='QC_data')
        repatch_data.to_excel(writer, sheet_name='repatch_all')
        juv_repatch.to_excel(writer, sheet_name='juvenile_data_repatch_QC')
        adult_repatch.to_excel(writer, sheet_name='adult_repatch_all_QC')

#analysis of intrinsic properties
def collect_connections_df(human_dir = '/Users/verjim/laptop_D_17.01.2022/Schmitz_lab/data/human/', 
                           folder = 'new'):
    exp_view = pd.read_excel(glob.glob(human_dir + '*experiments_overview.xlsx')[0])

    area_dict, high_k_dict = {}, {}
    for OP in exp_view['OP'].unique():
        area_dict[OP] = exp_view['region'][exp_view['OP'] == OP].tolist()[0]
        high_k_dict[OP] = exp_view['K concentration'][exp_view['OP'] == OP].tolist()[0]
    
    if folder != 'new':
        print('change the code to continue! sorry')

    connections_IC = glob.glob(human_dir + '/data_*/' + 'OP*' + '/connection_analysis/' + '*_connections.xlsx')
    all_cons_IC = pd.DataFrame()
    for IC_path in connections_IC:
        df = pd.read_excel(IC_path)
        if 'rosie' in IC_path:
            patcher = 'Rosie'
        else:
            patcher = 'Verji'
        df['pathcer'] = patcher
        all_cons_IC = pd.concat([all_cons_IC.loc[:], df]).reset_index(drop=True)
     
    all_cons_IC.insert(0, 'area', '')
    for OP in all_cons_IC['OP'].unique():
        mask = all_cons_IC['OP'] == OP
        if len(mask.unique()) == 1:
            continue
        all_cons_IC.loc[mask, 'area'] = area_dict[OP]
        all_cons_IC.loc[mask, 'high K concentration'] = high_k_dict[OP]
    all_cons_IC.loc[all_cons_IC['treatment'] == 'high k'] = 'high K'

    connections_VC = glob.glob(human_dir + '/data_*/' + 'OP*' + '/connection_analysis/' + '*_connected_cell_properties_post_in_VC.xlsx')
    all_cons_VC = pd.DataFrame()
    for VC_path in connections_VC:
        df = pd.read_excel(VC_path)
        all_cons_VC = pd.concat([all_cons_VC.loc[:], df]).reset_index(drop=True)
    
    all_cons_VC .insert(0, 'area', '')
    for OP in all_cons_VC ['OP'].unique():
        mask = all_cons_VC ['OP'] == OP
        if len(mask.unique()) == 1:
            continue
        all_cons_VC.loc[mask, 'area'] = area_dict[OP]
        all_cons_VC.loc[mask, 'high K concentration'] = high_k_dict[OP]
    all_cons_VC.loc[all_cons_VC ['treatment'] == 'high k'] = 'high K'

    date = str(datetime.date.today())

    file_name = '/Users/verjim/laptop_D_17.01.2022/Schmitz_lab/results/human/data/summary_data_tables/connectivity/' + date + '_connectivity.xlsx'
    with pd.ExcelWriter(file_name) as writer:
        all_cons_IC.to_excel(writer, sheet_name='Connectivity_IC')
        all_cons_VC.to_excel(writer, sheet_name='Connectivity_VC')

    return all_cons_IC, all_cons_VC

### Functions for fixing initial firing frequency tables


def create_IFF_data_table(OP, patcher, file_out = '_meta_active_chans.json', inj = 'full',
human_dir = '/Users/verjim/laptop_D_17.01.2022/Schmitz_lab/data/human/'):

    work_dir, filenames, indices_dict = sort.get_OP_metadata(human_dir, OP, patcher)[0:3]
    OP_meta = sort.from_json(work_dir, OP, file_out)

    treatments = OP_meta[0]['treatment']
    active_chans = OP_meta[0]['active_chans']
    slices = OP_meta[0]['slices']

    inj = hcf.read_inj(inj)
    col_names_inj = []
    for i in inj:
        col_names_inj.append(str(i) + 'pA')

    iterables = [col_names_inj, ["IFF", "num_aps"]]
    col_names = pd.MultiIndex.from_product(iterables, names=["current_inj", "parameter"])
    df_1 = pd.DataFrame(columns = ['OP', 'patcher', 'filename', 'slice', 'day', 'cell_ch', 'treatment'])
    df_2 = pd.DataFrame(columns = col_names)
    df_IFF = pd.concat([df_1, df_2], axis = 1)

    for i, char in enumerate(indices_dict['freq analyse']):
        fn = work_dir + filenames[char]
        channels = active_chans[i]
        treatment = treatments[i]
        if isinstance(treatment, list):
            treatment = treatments[i][0]
        slic = slices[i]
        day = 'D1'
        if slic[-2:] == 'D2':
            day = 'D2'

        param = hcf.get_initial_firing_rate(fn, channels, inj = 'full')

        df_OP1 = pd.DataFrame({'OP':OP,'patcher': patcher,'filename':filenames[char],'slice':slic,
        'day': day, 'cell_ch': channels, 'treatment': treatment})
        df_OP2 = pd.DataFrame(param, columns = col_names)
        df_IFF = pd.concat([df_IFF.loc[:], pd.concat([df_OP1, df_OP2], axis = 1)]).reset_index(drop=True)

    df_IFF.to_excel(work_dir + 'data_tables/' + OP + '_IFF_all.xlsx',index=False)

def collect_IFF_dfs(IFF_dirs = None ,human_dir = '/Users/verjim/laptop_D_17.01.2022/Schmitz_lab/data/human/'):
    exp_view = pd.read_excel(glob.glob(human_dir + '*experiments_overview.xlsx')[0])
    exp_view_IFF = exp_view[exp_view['Before and after'] == 'yes']

    area_dict, high_k_dict = {}, {}
    for OP in exp_view_IFF['OP'].unique():
        area_dict[OP] = exp_view_IFF['region'][exp_view_IFF['OP'] == OP].tolist()[0]
        high_k_dict[OP] = exp_view_IFF['K concentration'][exp_view_IFF['OP'] == OP].tolist()[0]
    
    IFF_dirs = glob.glob(human_dir + '/data_*/' + 'OP*' + '/data_tables/'  + '*_IFF*')
    IFF_all = pd.DataFrame()
    for IFF_path in IFF_dirs:
        df = pd.read_excel(IFF_path)
        IFF_all = pd.concat([IFF_all.loc[:], df]).reset_index(drop=True)
     
    IFF_all.insert(0, 'area', '')
    for OP in IFF_all['OP'].unique():
        mask = IFF_all['OP'] == OP
        IFF_all.loc[mask, 'area'] = area_dict[OP]
        IFF_all.loc[mask, 'high K concentration'] = high_k_dict[OP]

    mask = IFF_all['treatment'] == 'high k'
    IFF_all.loc[mask, 'treatment'] = 'high K'    #date = str(datetime.date.today())
    # IFF_all.to_excel('/Users/verjim/laptop_D_17.01.2022/Schmitz_lab/results/human/data/initial_firing_freqs/' 
    # + date + '_IFF_collected.xlsx', index=False)
    return IFF_all


def get_num_aps_and_IFF_data_culumns(df):
    num_aps_indx, IFF_indx = [], []
    for i in range(len(df.columns)):
        if 'num_aps' in df.columns[i]:
            num_aps_indx.append(i)
        if 'IFF' in df.columns[i]:
            IFF_indx.append(i)
    return num_aps_indx, IFF_indx


#### For manual EPSP analysis post-processing

def add_sheets_from_df_to_QC_checked_df():
    '''

    '''
    results_all_path = '/Users/verjim/laptop_D_17.01.2022/Schmitz_lab/data/human/meta_events/EPSPs/results_df/complete/'
    df_name = 'results_EPSPs_21.12.2023_manualy_modified.xlsx'

    results_df_QC = pd.read_excel(results_all_path[:-9] + 'QC_ed/QC_' + df_name)

    result_IDs_QC = []
    for j in range(len(results_df_QC)): #for each filename
        result_IDs_QC.append(results_df_QC['Recording filename'][j][:-4] + '_' + str(results_df_QC['Channel'][j]))

    results_df = load_workbook(results_all_path + df_name)
    old_file = load_workbook(results_all_path[:-9] + 'QC_ed/QC_' + df_name)

    writer = pd.ExcelWriter(results_all_path[:-9] + 'QC_ed/QC_' + df_name, engine = 'openpyxl')
    writer.book = old_file 

    for i in result_IDs_QC:
        df = pd.read_excel(results_all_path +  df_name, sheet_name = i)
        df.to_excel(writer, sheet_name = i)
    writer.close()

    # if you want to check if it worked
    # df_check = pd.ExcelFile(results_all_path[:-9] + 'QC_ed/QC_' + df_name)
    # missing_data = sorted(list(set(result_IDs_QC) - set(df_check.sheet_names)))
    # print(missing_data)

def check_if_data_missing_in_results_df():
    meta_dfs_dir = '/Users/verjim/laptop_D_17.01.2022/Schmitz_lab/data/human/meta_events/EPSPs/meta_dfs/original_/'
    meta_all = sort.concat_dfs_in_folder(meta_dfs_dir) 

    #find the duplicated rows
    duplicates_meta = meta_all.loc[meta_all.duplicated()].sort_values(['Name of recording', 'Channels to use']).reset_index(drop = True)

    meta_keep = meta_all.loc[meta_all['comment'] == 'keep'].reset_index(drop = True)

    meta_keep_IDs = []
    for i in range(len(meta_keep)):
        meta_keep_IDs.append(meta_keep['Name of recording'][i][:-4] + '_' + str(meta_keep['Channels to use'][i]))


    results_all_path = '/Users/verjim/laptop_D_17.01.2022/Schmitz_lab/data/human/meta_events/EPSPs/results_df/complete/'
    df_name = 'results_12.01.2024_all.xlsx'

    results_file = pd.ExcelFile(results_all_path + df_name)
    results_df = pd.read_excel(results_all_path +  df_name)

    result_IDs_all = []
    for j in range(len(results_df)):
        result_IDs_all.append(results_df['Recording filename'][j][:-4] + '_' + str(results_df['Channel'][j]))

    results_exclude = list(set(result_IDs_all) - set(meta_keep_IDs))

    if results_exclude != []:
        for k in results_exclude:
            fn = k[:-2] + '.abf'
            chan = np.int64(k[-1])
            indx = results_df.loc[(results_df['Recording filename'] == fn) & \
                (results_df['Channel'] == chan)].index
            results_df.drop(indx, axis=0, inplace=True)
        results_df.reset_index(inplace = True, drop = True)

    result_IDs_keep = []
    for g in range(len(results_df)):
        result_IDs_keep.append(results_df['Recording filename'][g][:-4] + '_' + str(results_df['Channel'][g]))

    missing_data = sorted(list(set(result_IDs_keep) - set(results_file.sheet_names)))
    print(missing_data)


def collect_events_dfs(event_type, human_dir = '/Users/verjim/laptop_D_17.01.2022/Schmitz_lab/data/human/'):
    ''' 
    event_type = 'minis' or 'spontan'
    '''
    OP_dirs = glob.glob(human_dir + 'data*' + '/' + 'OP*')

    meta_df = pd.DataFrame()
    for dir_ in OP_dirs:
        dir_ = dir_ + '/'

        df_QC_path = glob.glob(dir_ + '*data_tables/' + '*' + event_type + '.xlsx')
        if df_QC_path == []:
            continue
        QC_df = pd.read_excel(df_QC_path[0])
        if len(QC_df) == 0:
            continue
        patcher = QC_df['patcher'][0]
        OP = QC_df['OP'][0]

        meta_data_path = glob.glob(dir_ + '*data_tables/' + '*' + event_type + '_meta' + '*.xlsx')
        if meta_data_path == []:
            #create the metadata
            get_metadata_for_event_analysis(human_dir, OP, patcher, event_type)
            meta_data_path = glob.glob(dir_ + '*data_tables/' + '*' + event_type + '_meta' + '*.xlsx')
        df_meta = pd.read_excel(meta_data_path[0])
        df_meta = df_meta.loc[~df_meta['Name of recording'].isna()]
        meta_df = pd.concat([meta_df.loc[:], df_meta]).reset_index(drop=True)

    return meta_df

# collect minis
# fix or delete later
def wrong():
    exp_view_mini = exp_view[exp_view['minis'] == 'yes'].reset_index(drop=True)
    patcher_dict = {'Rosie': 'data_rosie/', 'Verji': 'data_verji/'}
    meta_df = pd.DataFrame()
    for i in range(len(exp_view_mini)): #ran to range 22 
        patcher =  exp_view_mini['patcher'][i]
        OP_dir  = human_dir + patcher_dict[patcher] + exp_view_mini['OP'][i] + '/data_tables/'
        QC_mini_path = glob.glob(OP_dir + '*' +'*final.xlsx')
        if len(QC_mini_path) > 0:
            df_QC_mini = pd.read_excel(QC_mini_path[0])
        else:
            print('No QC file found for ' + exp_view_mini['OP'][i])
            continue

        meta_df = pd.concat([meta_df[:], df_QC_mini]).reset_index(drop=True)
            
        # moving files directly to the recordings folder in event analysis
        for f in df_QC_mini.filename:
            shutil.copy(os.path.join(human_dir + patcher_dict[patcher] + exp_view_mini['OP'][i] + '/', f), \
                        '/Users/verjim/miniML_data/data/minis_events/')

    rows_to_delete = meta_df[meta_df.swps_to_analyse == '[]']

    # Delete the identified rows
    meta_df_cleaned = meta_df.drop(rows_to_delete.index)
    meta_df_cleaned.reset_index(drop=True, inplace=True)
    meta_df_cleaned.to_excel('/Users/verjim/laptop_D_17.01.2022/Schmitz_lab/data/human/meta_events/meta_files_to_analyse/' + date + 'minis_meta.xlsx')